from pathlib import Path
import csv
import re
import tarfile
import time
import zipfile
from io import BytesIO, TextIOWrapper

import orjson
from openpyxl import load_workbook
import pyarrow as pa
import pyarrow.parquet as pq
from sortedcontainers import SortedDict

from cex import cex_config
from cex.cex_common import build_part_path
from cex.cex_common import cleanup_stale_part_file
from cex.cex_common import download_file_from_storage
from cex.cex_common import list_storage_file_names
from cex.cex_common import replace_output_file
from cex.cex_common import seconds_until_next_utc_midnight
from cex.cex_common import storage_file_exists


BYBIT_DEPTH = 200  # Bybit历史订单簿深度，档位
BBO_DEPTH = 1  # BBO历史订单簿深度，档位
OKX_OUTPUT_DEPTHS = {
    "D10001": 400,  # OKX期货历史订单簿输出深度，档位
    "D10005": 400,  # OKX现货历史订单簿输出深度，档位
}  # OKX历史订单簿输出深度映射，映射
OKX_ORDERBOOK_LEVELS = {
    "D10001": "400lv",  # OKX期货历史订单簿档位标签，字符串
    "D10005": "400lv",  # OKX现货历史订单簿档位标签，字符串
}  # OKX历史订单簿档位标签映射，映射
DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")  # 日期格式正则，正则
BATCH_SIZE = 20000  # Parquet批次大小，条
SMALL_BATCH_SIZE = 100  # 超大深度Parquet批次大小，条
QUIET = False  # 静默模式开关，开关
STATUS_HOOK = None  # 状态回调函数，函数
LOG_HOOK = None  # 日志回调函数，函数
DATASET_QUIET = {}  # 分数据集静默模式映射，映射
DATASET_STATUS_HOOK = {}  # 分数据集状态回调映射，映射
DATASET_LOG_HOOK = {}  # 分数据集日志回调映射，映射


def configure_dataset_runtime(output_dataset_id: str, quiet: bool, status_hook, log_hook) -> None:
    """配置数据集运行时回调。"""
    DATASET_QUIET[output_dataset_id] = quiet
    DATASET_STATUS_HOOK[output_dataset_id] = status_hook
    DATASET_LOG_HOOK[output_dataset_id] = log_hook


def log(output_dataset_id: str, message: str) -> None:
    """输出指定数据集的日志消息。"""
    log_hook = DATASET_LOG_HOOK.get(output_dataset_id, LOG_HOOK)
    quiet = DATASET_QUIET.get(output_dataset_id, QUIET)
    if log_hook:
        log_hook(message)
    if not quiet:
        print(message)


def status_update(output_dataset_id: str, exchange: str, symbol: str, value) -> None:
    """更新指定数据集的状态信息。"""
    hook = DATASET_STATUS_HOOK.get(output_dataset_id, STATUS_HOOK)
    if not hook:
        return
    market = "future" if output_dataset_id == "D10011" else "spot"
    hook(cex_config.get_status_key(exchange, market, symbol), value)


def list_input_symbols(base_dir: Path) -> list[str]:
    """列出输入目录中的交易对目录。"""
    if not base_dir.exists():
        return []
    return sorted([path.name for path in base_dir.iterdir() if path.is_dir()])


def resolve_symbols(input_dataset_id: str, exchange: str, base_dir: Path) -> list[str]:
    """解析需要处理的交易对列表。"""
    if input_dataset_id == "D10001":
        configured = cex_config.get_future_symbols(exchange)
    else:
        configured = cex_config.get_spot_symbols(exchange)
    return sorted(set(configured) | set(list_input_symbols(base_dir)))


def output_depth_for_dataset(input_dataset_id: str, exchange: str) -> int:
    """返回数据集对应的输出深度。"""
    if exchange == "bybit":
        return BYBIT_DEPTH
    if exchange in {"binance", "bitget"}:
        return BBO_DEPTH
    return OKX_OUTPUT_DEPTHS[input_dataset_id]


def okx_orderbook_level_for_dataset(input_dataset_id: str) -> str:
    """返回OKX数据集对应的订单簿档位标签。"""
    return OKX_ORDERBOOK_LEVELS[input_dataset_id]


def batch_size_for_dataset(input_dataset_id: str, exchange: str) -> int:
    """返回数据集对应的Parquet批次大小。"""
    if exchange == "okx" and output_depth_for_dataset(input_dataset_id, exchange) > 1000:
        return SMALL_BATCH_SIZE
    return BATCH_SIZE


def build_input_path(input_dataset_id: str, exchange: str, base_dir: Path, symbol: str, date_str: str) -> Path:
    """构造输入归档路径。"""
    if exchange == "bybit":
        return base_dir / symbol / f"{date_str}_{symbol}_ob{BYBIT_DEPTH}.data.zip"
    if exchange == "binance":
        return base_dir / symbol / f"{symbol}-bookTicker-{date_str}.zip"
    if exchange == "bitget":
        return base_dir / symbol / f"{date_str.replace('-', '')}.zip"
    return base_dir / symbol / f"{symbol}-L2orderbook-{okx_orderbook_level_for_dataset(input_dataset_id)}-{date_str}.tar.gz"


def build_output_path(input_dataset_id: str, exchange: str, base_dir: Path, symbol: str, date_str: str) -> Path:
    """构造输出快照路径。"""
    date_tag = date_str.replace("-", "")
    depth = output_depth_for_dataset(input_dataset_id, exchange)
    return base_dir / symbol / date_tag / f"{date_tag}_{symbol}_ob{depth}_snapshot.parquet"


def iter_zip_messages(file_path: Path):
    """遍历Zip文件中的原始订单簿消息。"""
    with zipfile.ZipFile(file_path, "r") as zip_file:
        for name in zip_file.namelist():
            with zip_file.open(name) as file_obj:
                for line in file_obj:
                    yield orjson.loads(line)


def iter_binance_messages(file_path: Path, symbol: str):
    """遍历Binance原始BBO消息。"""
    with zipfile.ZipFile(file_path, "r") as zip_file:
        name = zip_file.namelist()[0]
        with zip_file.open(name) as file_obj:
            reader = csv.DictReader(TextIOWrapper(file_obj, encoding="utf-8", newline=""))
            for row in reader:
                row["symbol"] = symbol
                yield row


def iter_bitget_messages(file_path: Path, symbol: str):
    """遍历Bitget原始BBO消息。"""
    with zipfile.ZipFile(file_path, "r") as zip_file:
        name = zip_file.namelist()[0]
        workbook = load_workbook(BytesIO(zip_file.read(name)), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = [str(item) for item in next(rows)]
        for values in rows:
            row = {header[index]: values[index] for index in range(len(header))}
            row["symbol"] = symbol
            yield row


def iter_tar_messages(file_path: Path):
    """遍历Tar文件中的原始订单簿消息。"""
    with tarfile.open(file_path, "r:gz") as tar_file:
        for member in tar_file.getmembers():
            if not member.isfile():
                continue
            file_obj = tar_file.extractfile(member)
            if not file_obj:
                continue
            for line in file_obj:
                yield orjson.loads(line)


def iter_messages(exchange: str, file_path: Path, symbol: str):
    """遍历归档文件中的原始订单簿消息。"""
    if exchange == "bybit":
        yield from iter_zip_messages(file_path)
        return
    if exchange == "binance":
        yield from iter_binance_messages(file_path, symbol)
        return
    if exchange == "bitget":
        yield from iter_bitget_messages(file_path, symbol)
        return
    yield from iter_tar_messages(file_path)


def trim_levels(levels: list) -> list[list[str]]:
    """裁剪盘口层级为价格和数量。"""
    return [[str(level[0]), str(level[1])] for level in levels if len(level) >= 2]


def normalize_ts_ms(value) -> int | None:
    """将不同精度时间戳统一为毫秒。"""
    parsed = parse_int_or_none(value)
    if parsed is None:
        return None
    if abs(parsed) < 10**11:
        return parsed * 1000
    return parsed


def normalize_message(exchange: str, msg: dict) -> tuple[str, dict] | None:
    """归一化不同交易所的原始消息结构。"""
    if exchange == "bybit":
        return msg.get("type", ""), msg.get("data", {})
    if exchange == "binance":
        return (
            "snapshot",
            {
                "s": msg.get("symbol"),
                "b": [[msg.get("best_bid_price"), msg.get("best_bid_qty")]],
                "a": [[msg.get("best_ask_price"), msg.get("best_ask_qty")]],
                "u": msg.get("update_id"),
            },
        )
    if exchange == "bitget":
        return (
            "snapshot",
            {
                "s": msg.get("symbol"),
                "b": [[msg.get("bid_price"), msg.get("bid_volume")]],
                "a": [[msg.get("ask_price"), msg.get("ask_volume")]],
            },
        )
    action = msg.get("action", "")
    if action not in {"snapshot", "update"}:
        return None
    return (
        "snapshot" if action == "snapshot" else "delta",
        {
            "s": msg.get("instId"),
            "b": trim_levels(msg.get("bids", [])),
            "a": trim_levels(msg.get("asks", [])),
        },
    )


def update_orderbook(orderbook: dict, msg_type: str, data: dict) -> bool:
    """将原始消息应用到内存盘口。"""
    if msg_type == "snapshot":
        orderbook["bids"] = SortedDict({float(bid[0]): float(bid[1]) for bid in data.get("b", [])})
        orderbook["asks"] = SortedDict({float(ask[0]): float(ask[1]) for ask in data.get("a", [])})
        return True
    if msg_type == "delta":
        for price_text, size_text in data.get("b", []):
            price = float(price_text)
            size = float(size_text)
            if size == 0:
                orderbook["bids"].pop(price, None)
            else:
                orderbook["bids"][price] = size
        for price_text, size_text in data.get("a", []):
            price = float(price_text)
            size = float(size_text)
            if size == 0:
                orderbook["asks"].pop(price, None)
            else:
                orderbook["asks"][price] = size
        return True
    return False


def parse_int_or_none(value) -> int | None:
    """将数值安全转换为整数。"""
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value)
    if text.startswith("-"):
        return -int(text[1:]) if text[1:].isdigit() else None
    return int(text) if text.isdigit() else None


def build_snapshot(input_dataset_id: str, exchange: str, orderbook: dict, msg: dict, msg_type: str, data: dict) -> dict:
    """构造Parquet写入快照。"""
    depth = output_depth_for_dataset(input_dataset_id, exchange)
    bids = orderbook["bids"]
    asks = orderbook["asks"]
    bid_items = list(reversed(bids.items()))[:depth]
    ask_items = list(asks.items())[:depth]
    best_bid = bid_items[0][0] if bid_items else None
    best_ask = ask_items[0][0] if ask_items else None
    bid_list = [{"price": price, "qty": qty} for price, qty in bid_items]
    ask_list = [{"price": price, "qty": qty} for price, qty in ask_items]
    ts_value = normalize_ts_ms(msg.get("ts"))
    cts_value = normalize_ts_ms(msg.get("cts"))
    return {
        "symbol": data.get("s", msg.get("symbol")),
        "update_type": msg_type,
        "ts": ts_value,
        "cts": cts_value if cts_value is not None else ts_value,
        "update_id": parse_int_or_none(data.get("u")),
        "seq": parse_int_or_none(data.get("seq")),
        "best_bid": best_bid,
        "best_ask": best_ask,
        "bid_depth": len(bid_items),
        "ask_depth": len(ask_items),
        "bids": bid_list,
        "asks": ask_list,
    }


def build_schema() -> pa.Schema:
    """构造输出Parquet表结构。"""
    side_type = pa.list_(pa.struct([("price", pa.float64()), ("qty", pa.float64())]))
    return pa.schema(
        [
            ("symbol", pa.string()),
            ("update_type", pa.string()),
            ("ts", pa.int64()),
            ("cts", pa.int64()),
            ("update_id", pa.int64()),
            ("seq", pa.int64()),
            ("best_bid", pa.float64()),
            ("best_ask", pa.float64()),
            ("bid_depth", pa.int32()),
            ("ask_depth", pa.int32()),
            ("bids", side_type),
            ("asks", side_type),
        ]
    )


def write_parquet(records: list, writer: pq.ParquetWriter, schema: pa.Schema) -> None:
    """写入单个Parquet批次。"""
    writer.write_table(pa.Table.from_pylist(records, schema=schema))


def is_valid_archive(file_path: Path) -> bool:
    """判断归档文件是否为有效压缩包。"""
    if file_path.name.endswith(".zip"):
        return zipfile.is_zipfile(file_path)
    if file_path.name.endswith(".tar.gz"):
        return tarfile.is_tarfile(file_path)
    return False


def process_date(input_dataset_id: str, output_dataset_id: str, exchange: str, input_dir: Path, output_dir: Path, symbol: str, date_str: str) -> None:
    """处理单日订单簿归档。"""
    input_path = build_input_path(input_dataset_id, exchange, input_dir, symbol, date_str)
    if not input_path.exists() and not download_file_from_storage(input_path):
        return
    if not is_valid_archive(input_path):
        log(output_dataset_id, f"文件不是有效压缩包: {input_path}")
        input_path.unlink()
        return
    output_path = build_output_path(input_dataset_id, exchange, output_dir, symbol, date_str)
    cleanup_stale_part_file(output_path)
    if storage_file_exists(output_path):
        return
    tmp_output_path = build_part_path(output_path)
    if tmp_output_path.exists():
        tmp_output_path.unlink()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    schema = build_schema()
    writer = pq.ParquetWriter(tmp_output_path, schema, compression="snappy")
    orderbook = {"bids": SortedDict(), "asks": SortedDict()}
    has_snapshot = False
    batch = []
    batch_size = batch_size_for_dataset(input_dataset_id, exchange)
    total = 0
    for msg in iter_messages(exchange, input_path, symbol):
        normalized = normalize_message(exchange, msg)
        if normalized is None:
            continue
        msg_type, data = normalized
        if msg_type == "snapshot":
            has_snapshot = update_orderbook(orderbook, msg_type, data)
        elif msg_type == "delta" and has_snapshot:
            update_orderbook(orderbook, msg_type, data)
        else:
            continue
        batch.append(build_snapshot(input_dataset_id, exchange, orderbook, msg, msg_type, data))
        total += 1
        if len(batch) >= batch_size:
            write_parquet(batch, writer, schema)
            batch.clear()
    if batch:
        write_parquet(batch, writer, schema)
    writer.close()
    replace_output_file(tmp_output_path, output_path)
    log(output_dataset_id, f"已写入: {output_path}，记录数: {total}")


def parse_date_from_name(input_dataset_id: str, exchange: str, file_name: str, symbol: str) -> str | None:
    """从归档文件名中解析日期。"""
    if exchange == "bybit":
        suffix = f"_{symbol}_ob{BYBIT_DEPTH}.data.zip"
        if not file_name.endswith(suffix):
            return None
        date_text = file_name[: -len(suffix)]
        return date_text if DATE_PATTERN.fullmatch(date_text) else None
    if exchange == "binance":
        prefix = f"{symbol}-bookTicker-"
        suffix = ".zip"
        if not file_name.startswith(prefix) or not file_name.endswith(suffix):
            return None
        date_text = file_name[len(prefix) : -len(suffix)]
        return date_text if DATE_PATTERN.fullmatch(date_text) else None
    if exchange == "bitget":
        if not file_name.endswith(".zip"):
            return None
        raw_date = file_name.removesuffix(".zip")
        if len(raw_date) != 8 or not raw_date.isdigit():
            return None
        date_text = f"{raw_date[0:4]}-{raw_date[4:6]}-{raw_date[6:8]}"
        return date_text if DATE_PATTERN.fullmatch(date_text) else None
    prefix = f"{symbol}-L2orderbook-{okx_orderbook_level_for_dataset(input_dataset_id)}-"
    suffix = ".tar.gz"
    if not file_name.startswith(prefix) or not file_name.endswith(suffix):
        return None
    date_text = file_name[len(prefix) : -len(suffix)]
    return date_text if DATE_PATTERN.fullmatch(date_text) else None


def processed_dates_for_symbol(
    input_dataset_id: str,
    exchange: str,
    output_dir: Path,
    symbol: str,
    candidate_dates: list[str],
) -> set[str]:
    """统计指定交易对已处理完成的日期集合。"""
    processed = set()
    for date_str in candidate_dates:
        output_path = build_output_path(input_dataset_id, exchange, output_dir, symbol, date_str)
        if storage_file_exists(output_path):
            processed.add(date_str)
    return processed


def iter_available_dates(input_dataset_id: str, exchange: str, input_dir: Path, symbol: str, start_date: str) -> list[str]:
    """遍历某个交易对可处理的日期。"""
    dates = set()
    for file_name in list_storage_file_names(input_dir / symbol):
        date_text = parse_date_from_name(input_dataset_id, exchange, file_name, symbol)
        if date_text and date_text >= start_date:
            dates.add(date_text)
    return sorted(dates)


def run_dataset(input_dataset_id: str, output_dataset_id: str) -> None:
    """运行指定历史订单簿快照任务。"""
    while True:
        for exchange in cex_config.get_supported_exchanges(output_dataset_id):
            input_dir = cex_config.get_source_dir(input_dataset_id, exchange)
            output_dir = cex_config.get_output_dir(output_dataset_id, exchange)
            start_date = cex_config.get_min_start_date(input_dataset_id, exchange)
            if not input_dir or not output_dir or not start_date:
                continue
            for symbol in resolve_symbols(input_dataset_id, exchange, input_dir):
                available_dates = iter_available_dates(input_dataset_id, exchange, input_dir, symbol, start_date)
                if not available_dates:
                    status_update(output_dataset_id, exchange, symbol, (0, "无可处理数据"))
                    continue
                processed_dates = processed_dates_for_symbol(input_dataset_id, exchange, output_dir, symbol, available_dates)
                done_count = len(processed_dates)
                synced_until = cex_config.get_min_start_date(input_dataset_id, exchange)
                if processed_dates:
                    synced_until = max(processed_dates)
                    status_update(output_dataset_id, exchange, symbol, (done_count, f"日 {synced_until} 准备回补"))
                else:
                    status_update(output_dataset_id, exchange, symbol, (0, f"准备 {available_dates[0]}"))
                for date_str in available_dates:
                    if date_str in processed_dates:
                        continue
                    status_update(output_dataset_id, exchange, symbol, (done_count, f"日 {date_str} 请求中"))
                    process_date(input_dataset_id, output_dataset_id, exchange, input_dir, output_dir, symbol, date_str)
                    done_count += 1
                    output_name = build_output_path(input_dataset_id, exchange, output_dir, symbol, date_str).name
                    status_update(output_dataset_id, exchange, symbol, (done_count, f"日 {date_str} {output_name}"))
        sleep_seconds = seconds_until_next_utc_midnight()
        log(output_dataset_id, f"等待 {sleep_seconds} 秒后再次执行（UTC 00:00）")
        time.sleep(sleep_seconds)
